"""Utility helpers for generating and managing template Excel files."""

from __future__ import annotations

from dataclasses import dataclass
import re
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Sequence
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from app.domain.entities import TemplateColumn
from app.infrastructure.storage import delete_blob, download_blob_to_path, upload_blob

_TEMPLATE_PREFIX = "Templates"
_EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_DOWNLOAD_DIRECTORY = Path(tempfile.gettempdir()) / "accura_api_templates"


@dataclass
class TemplateExcelInfo:
    """Metadata returned after uploading a template workbook."""

    filename: str
    blob_path: str
    size_bytes: int


def _sanitize_filename(name: str) -> str:
    """Return ``name`` transformed into a filesystem-safe slug."""

    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", name.strip())
    cleaned = cleaned.strip("_")
    return cleaned or "template"


def _build_blob_path(
    template_id: int, user_id: int, table_name: str, filename: str
) -> str:
    return f"{_TEMPLATE_PREFIX}/{template_id}-{user_id}-{table_name}-{filename}"


def _create_workbook(columns: Sequence[TemplateColumn]) -> Workbook:
    workbook = Workbook()
    workbook.security.lockStructure = True
    worksheet = workbook.active
    worksheet.title = "Datos"

    headers = [column.name for column in columns]
    if headers:
        worksheet.append(headers)
        header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.protection = Protection(locked=True)

        worksheet.freeze_panes = "A2"

        max_col = len(headers)
        for row in worksheet.iter_rows(min_row=2, max_row=1000, max_col=max_col):
            for cell in row:
                cell.protection = Protection(locked=False)

    worksheet.protection.enable()
    worksheet.protection.insertColumns = False
    worksheet.protection.deleteColumns = False
    worksheet.protection.insertRows = True
    worksheet.protection.deleteRows = False
    worksheet.protection.formatColumns = False
    worksheet.protection.formatRows = False
    worksheet.protection.sort = False
    worksheet.protection.autoFilter = False
    worksheet.protection.insertHyperlinks = False

    return workbook


def create_template_excel(
    template_id: int,
    template_name: str,
    columns: Sequence[TemplateColumn],
    *,
    user_id: int,
    table_name: str,
) -> TemplateExcelInfo:
    """Create an Excel workbook for ``template_id`` and upload it to Blob Storage."""

    workbook = _create_workbook(columns)
    buffer = BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()

    filename = f"{template_id}_{_sanitize_filename(template_name)}.xlsx"
    blob_path = _build_blob_path(template_id, user_id, table_name, filename)
    upload_blob(blob_path, data, content_type=_EXCEL_CONTENT_TYPE)

    return TemplateExcelInfo(
        filename=filename,
        blob_path=blob_path,
        size_bytes=len(data),
    )


def delete_template_excel(blob_path: str) -> None:
    """Delete the template Excel file stored at ``blob_path``."""

    delete_blob(blob_path)


def download_template_excel(blob_path: str) -> Path:
    """Download the template Excel stored at ``blob_path`` to a temp file."""

    if not blob_path:
        raise FileNotFoundError("Ruta de plantilla no disponible")

    _DOWNLOAD_DIRECTORY.mkdir(parents=True, exist_ok=True)
    destination = _DOWNLOAD_DIRECTORY / f"{uuid4().hex}.xlsx"
    return download_blob_to_path(blob_path, destination)


__all__ = [
    "TemplateExcelInfo",
    "create_template_excel",
    "delete_template_excel",
    "download_template_excel",
]

